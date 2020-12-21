from openpyxl import Workbook
from datetime import datetime

def make_excel(y_pos1, y_pos2, y_pos3, y_pos4, y_pos5, y_pos6):
    # 엑셀 파일 저장 위치
    excel_file_path = 'C:/Users/battl/PycharmProjects/ComputerScienceEngineering/project list/Optimum Capacity Design of Renewable Energy Source Project/excel/'
    # 엑셀 파일 이름
    excel_file_name = excel_file_path + 'simulation.xlsx'
    # 엑셀 sheet 이름
    excel_sheet_title = '1'

    work_book = Workbook()
    sheet1 = work_book.active
    sheet1.title = excel_sheet_title

    # 헤더 입력
    sheet1.cell(row=1, column=1).value = 'Pwt'
    sheet1.cell(row=1, column=2).value = 'Pdummy'
    sheet1.cell(row=1, column=3).value = 'Ppv'
    sheet1.cell(row=1, column=4).value = 'Pbat'
    sheet1.cell(row=1, column=5).value = 'Pload'
    sheet1.cell(row=1, column=6).value = 'SOC'

    # 엑셀 행, 열
    excel_row = 2
    excel_column = 1
    for idx, values in enumerate(y_pos1):
        sheet1.cell(row=excel_row, column=excel_column).value = values
        excel_row += 1

    # 엑셀 행, 열
    excel_row = 2
    excel_column = 2
    for idx, values in enumerate(y_pos2):
        sheet1.cell(row=excel_row, column=excel_column).value = values
        excel_row += 1

    # 엑셀 행, 열
    excel_row = 2
    excel_column = 3
    for idx, values in enumerate(y_pos3):
        sheet1.cell(row=excel_row, column=excel_column).value = values
        excel_row += 1

    # 엑셀 행, 열
    excel_row = 2
    excel_column = 4
    for idx, values in enumerate(y_pos4):
        sheet1.cell(row=excel_row, column=excel_column).value = values
        excel_row += 1

    # 엑셀 행, 열
    excel_row = 2
    excel_column = 5
    for idx, values in enumerate(y_pos5):
        sheet1.cell(row=excel_row, column=excel_column).value = values
        excel_row += 1

    # 엑셀 행, 열
    excel_row = 2
    excel_column = 6
    for idx, values in enumerate(y_pos6):
        sheet1.cell(row=excel_row, column=excel_column).value = values
        excel_row += 1

    # 엑셀 파일 크기 조정하기
    work_book.save(filename=excel_file_name)
    work_book.close()