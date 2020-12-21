from matplotlib import pyplot as plt
from openpyxl import Workbook

# Enter capacity(kW) and daily pattern data of WT, PV, Load
Pwt_arg = 5
Pwt = [
    0.81, 1.00, 0.57, 0.38, 0.22, 0.35, 0.44, 0.36,
    0.57, 0.82, 0.58, 0.33, 0.35, 0.64, 0.64, 0.22,
    0.18, 0.10, 0.05, 0.08, 0.15, 0.08, 0.14, 0.13
]

Ppv_arg = 10
Ppv = [
    0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0,
    0.2, 0.5, 0.7, 0.9, 1.0, 0.9, 0.8, 0.7,
    0.53125, 0.2, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
]

Pload_arg = 10
Pload = [
    0.33, 0.32, 0.32, 0.31, 0.30, 0.30, 0.34, 0.40,
    0.50, 0.70, 0.69, 0.60, 0.55, 0.70, 0.68, 0.60,
    0.53, 0.50, 0.47, 0.45, 0.40, 0.36, 0.34, 0.33
]

# Enter efficiency of inverter
Ef_inv = 0.95

# Enter constraints of LOLP and dummy loads
Con_LOLP = 0.05
Con_dummy = 0.04

# Enter battery data, initial SOC(%), SOC_min(%), SOC_max(%), P_bc(kW), P_bd(kW)
C_bat = 20				# Capacity(kWh)
SOC_ini = 30			# initial SOC of battery
SOC_min = 10			# minimum SOC of battery
SOC_max = 90			# maximum SOC of battery
P_bc = [0] * 24		    # output of battery in charge mode
P_bd = [0] * 24		    # output of battery in discharge mode
Ef_bc = 0.9			    # efficiency of battery in charge mode
Ef_bd = 0.85			# efficiency of battery in discharge mode
r_sd = 0.002			# self - discharge rate of battery

# Enter dummy load capacity(kW)
P_dummy = [0] * 24
E_dummy = 0

# Enter diesel generator capacity(kW), output(kW) and efficiency
P_dgr = 10
P_dg = [0] * 24
Ef_dg = 0.95

# NPSP algorithm
t = 1							# initial time
LOLP = [0] * 24					# initial LOLP at each time
LOLP_sum = 0
E_bat = (C_bat * SOC_ini) / 100	# initial kWh capacity of battery at each time
SOC = [0] * 24					# initial SOC of battery at each time

# iteration for day calculation(24h)
for t in range(24):
    Pwt[t] = Pwt_arg * Pwt[t]
    Ppv[t] = Ppv_arg * Ppv[t]
    Pload[t] = Pload_arg * Pload[t]

    SOC[t] = E_bat / C_bat * 100		# SOC calculation of battery at each time

    # occurred surplus energy in HRES system
    if Pwt[t] > Pload[t]:
        # check SOC of battery whether the SOC has enough
        if SOC[t] < SOC_max:
            P_bc[t] = (Pwt[t] - Pload[t] * Ef_inv + Ppv[t]) * Ef_bc		    # calculate charge capacity of battery
            E_bat = E_bat * (1 - r_sd) + P_bc[t] * Ef_bc				    #  accumulate capacity of battery
        else:
            P_bc[t] = 0													    # calculate charge capacity of battery
            P_dummy[t] = (Pwt[t] - Pload[t]) + Ppv[t] * Ef_inv			    # calculate capacity of dummy loads to cover surplus energy in HRES system
            E_dummy = E_dummy + P_dummy[t]								    # accumulate capacity of dummy loads to check total used capacity of dummy loads

    # occurred surplus energy in HRES system
    if Pwt[t] <= Pload[t] and (Pwt[t] + Ppv[t] * Ef_inv) > Pload[t]:
        # check SOC of battery whether the SOC has enough
        if SOC[t] < SOC_max:
            P_bc[t] = (Pwt[t] + Ppv[t] * Ef_inv - Pload[t]) * Ef_bc*Ef_inv  # calculate charge capacity of battery
            E_bat = E_bat * (1 - r_sd) + P_bc[t] * Ef_bc					# accumulate capacity of battery
        else:
            P_bc[t] = 0													    # calculate charge capacity of battery
            P_dummy[t] = Pwt[t] + Ppv[t]*Ef_inv - Pload[t]					# calculate capacity of dummy loads to cover surplus energy in HRES system
            E_dummy = E_dummy + P_dummy[t]									# accumulate capacity of dummy loads to check total used capacity of dummy loads

    # occurred deficit energy in HRES system
    if Pwt[t] <= Pload[t] and (Pwt[t] + Ppv[t]*Ef_inv) <= Pload[t]:
        # check SOC of battery whether the SOC has lack of capacity
        if SOC[t] > SOC_min:
            P_bd[t] = (Ppv[t] - (Pload[t] - Pwt[t]) / Ef_inv) * Ef_bd		# calculate discharge capacity of battery
            E_bat = E_bat * (1 - r_sd) - abs(P_bd[t]) * Ef_bd				# accumulate capacity of battery
        else:
            P_bd[t] = 0													    # calculate charge capacity of battery
            P_dg[t] = (Pload[t] - Pwt[t] - Ppv[t] * Ef_inv) / Ef_dg		    # calculate capacity of diesel generator to cover deficit energy in HRES system
            # check the capacity of diesel generator whether it exceeds rated capacity of diesel generator
            if P_dg[t] > P_dgr:
                LOLP[t] = 1					    # evaluate LOLP at time t
                LOLP_sum = LOLP_sum + LOLP[t]   # accumulate LOLP at each time
            else:
                LOLP[t] = 0				        # evaluate LOLP at time t

    # check capacity of HRES with constraints of LOLP and dummy loads
    if E_dummy <= 0: print("Need to increasing capacity of PV or WT.\n")
    else:
        if E_dummy > max(Pload): print("Need to decreasing capacity of PV or WT.\n")
        else:
            if LOLP_sum / t > Con_LOLP: print("Need to increasing capacity of PV or WT.\n")
            else: print("Present capacities of PV and WT are optimal.\n")

# Pwt 그래프
plt.subplot(3, 2, 1)
x_pos = [x for x in range(24)]
y_pos1 = [round(Pwt[idx], 3) for idx, y in enumerate(Pwt)]
plt.plot(x_pos, y_pos1)
plt.xlabel('t(h)')
plt.ylabel('Pwt(kWh)')

# Pdummy 그래프
plt.subplot(3, 2, 2)
y_pos2 = [round(P_dummy[idx], 3) for idx, y in enumerate(P_dummy)]
plt.plot(x_pos, y_pos2)
plt.xlabel('t(h)')
plt.ylabel('Pdummy(kWh)')

# Ppv 그래프
plt.subplot(3, 2, 3)
y_pos3 = [round(Ppv[idx], 3) for idx, y in enumerate(Ppv)]
plt.plot(x_pos, y_pos3)
plt.xlabel('t(h)')
plt.ylabel('Ppv(kWh)')

# Pbat 그래프
plt.subplot(3, 2, 4)
y_pos4 = [round(bc + bd, 3) for bc, bd in zip(P_bc, P_bd)]
plt.plot(x_pos, y_pos4)
plt.xlabel('t(h)')
plt.ylabel('Pbat(kWh)')

# Pload 그래프
plt.subplot(3, 2, 5)
y_pos5 = [round(Pload[idx], 3) for idx, y in enumerate(Pload)]
plt.plot(x_pos, y_pos5)
plt.xlabel('t(h)')
plt.ylabel('Pload(kWh)')

# SOC 그래프
plt.subplot(3, 2, 6)
y_pos6 = [round(SOC[idx], 3) for idx, y in enumerate(SOC)]
plt.plot(x_pos, y_pos6)
plt.xlabel('t(h)')
plt.ylabel('SOC(%)')

plt.show()



# 엑셀 파일 저장 위치
excel_file_path = 'C:/Users/battl/PycharmProjects/ComputerScienceEngineering/project list/Optimum Capacity Design of Renewable Energy Source Project/excel/'
# 엑셀 파일 이름
excel_file_name = 'simulation.xlsx'
# 엑셀 sheet 이름
excel_sheet_title = '1'

work_book = Workbook()
sheet1 = work_book.active
sheet1.title = excel_sheet_title

# 헤더 입력
sheet1.cell(row=1, column=1).value = 'Pwt'
sheet1.cell(row=1, column=2).value = 'Pdummy'
sheet1.cell(row=1, column=3).value = 'Ppv'
sheet1.cell(row=1, column=4).value = 'Pbat'
sheet1.cell(row=1, column=5).value = 'Pload'
sheet1.cell(row=1, column=6).value = 'SOC'

# 엑셀 행, 열
excel_row = 2
excel_column = 1
for idx, values in enumerate(y_pos1):
    sheet1.cell(row=excel_row, column=excel_column).value = values
    excel_row += 1

# 엑셀 행, 열
excel_row = 2
excel_column = 2
for idx, values in enumerate(y_pos2):
    sheet1.cell(row=excel_row, column=excel_column).value = values
    excel_row += 1

# 엑셀 행, 열
excel_row = 2
excel_column = 3
for idx, values in enumerate(y_pos3):
    sheet1.cell(row=excel_row, column=excel_column).value = values
    excel_row += 1

# 엑셀 행, 열
excel_row = 2
excel_column = 4
for idx, values in enumerate(y_pos4):
    sheet1.cell(row=excel_row, column=excel_column).value = values
    excel_row += 1

# 엑셀 행, 열
excel_row = 2
excel_column = 5
for idx, values in enumerate(y_pos5):
    sheet1.cell(row=excel_row, column=excel_column).value = values
    excel_row += 1

# 엑셀 행, 열
excel_row = 2
excel_column = 6
for idx, values in enumerate(y_pos6):
    sheet1.cell(row=excel_row, column=excel_column).value = values
    excel_row += 1

# 엑셀 파일 크기 조정하기
work_book.save(filename=excel_file_name)
work_book.close()